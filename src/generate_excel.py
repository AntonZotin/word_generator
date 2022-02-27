from openpyxl import load_workbook
from datetime import datetime, date
from src.strings import XLSX_FILE_SUFFIX, XLSX_FILE_PREFIX, XLSX_NAME, XLSX_PP, XLSX_INN, XLSX_NUMBER, \
    XLSX_ISPOLNITEL, XLSX_DATE


def insert_values_in_row(ws, values, row):
    ws[f'{XLSX_NAME}{row}'].value = values[0]
    ws[f'{XLSX_INN}{row}'].value = values[1]
    ws[f'{XLSX_NUMBER}{row}'].value = values[2]
    ws[f'{XLSX_DATE}{row}'].value = values[3]
    ws[f'{XLSX_ISPOLNITEL}{row}'].value = values[4]


def get_str_date(ws, row):
    return str(ws[f'{XLSX_DATE}{row.row}'].value.strftime("%d.%m.%Y")) \
        if isinstance(ws[f'{XLSX_DATE}{row.row}'].value, datetime) \
           or isinstance(ws[f'{XLSX_DATE}{row.row}'].value, date) else ws[f'{XLSX_DATE}{row.row}'].value


def main_insert_and_sort_xlsx(data):
    number_post = '327' if postanovlenie == 'Процентная ставка' else '326'
    number_prefix = 'П' if postanovlenie == 'Процентная ставка' else 'Д'
    document_name = f'../{XLSX_FILE_PREFIX}{number_post}{XLSX_FILE_SUFFIX}'
    document = load_workbook(document_name)
    ws = document.active
    new_row = [name, inn, f'{number_prefix}-{number}', date, ispolnitel]
    rows = [new_row, ]
    begin_row = 1
    for row in ws[XLSX_NAME]:
        if row.row == 1:
            continue
        elif not row.value:
            if ws[f'{XLSX_PP}{row.row}'].value:
                break
            else:
                rows = [new_row, ]
                begin_row = row.row
                continue
        else:
            rows.append([
                ws[f'{XLSX_NAME}{row.row}'].value,
                ws[f'{XLSX_INN}{row.row}'].value,
                ws[f'{XLSX_NUMBER}{row.row}'].value,
                get_str_date(ws, row),
                ws[f'{XLSX_ISPOLNITEL}{row.row}'].value
            ])
    sorted_rows = sorted(rows, key=lambda r: r[2])
    for row in range(len(sorted_rows)):
        insert_values_in_row(ws, sorted_rows[row], begin_row + row + 1)
    document.save(document_name)
