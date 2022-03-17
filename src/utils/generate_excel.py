from datetime import datetime, date

from openpyxl import load_workbook

from src.utils.strings import XLSX_FILE_SUFFIX, XLSX_NAME, XLSX_PP, XLSX_INN, XLSX_NUMBER, \
    XLSX_ISPOLNITEL, XLSX_REQUEST_DATE, XLSX_COMMENT, XLSX_SUMM, XLSX_YANDEX, XLSX_DELIVERY, XLSX_CHECK_DATE, \
    XLSX_FILE_EDA, XLSX_FILE_STAVKA


def insert_values_in_row(ws, values, row):
    ws[f'{XLSX_PP}{row}'].value = values[0]
    ws[f'{XLSX_NAME}{row}'].value = values[1]
    ws[f'{XLSX_INN}{row}'].value = values[2]
    ws[f'{XLSX_NUMBER}{row}'].value = values[3]
    ws[f'{XLSX_REQUEST_DATE}{row}'].value = values[4]
    ws[f'{XLSX_COMMENT}{row}'].value = values[5]
    ws[f'{XLSX_SUMM}{row}'].value = values[6]
    ws[f'{XLSX_YANDEX}{row}'].value = values[7]
    ws[f'{XLSX_DELIVERY}{row}'].value = values[8]
    ws[f'{XLSX_ISPOLNITEL}{row}'].value = values[9]
    ws[f'{XLSX_CHECK_DATE}{row}'].value = values[10]


def get_str_date(ws, row):
    return str(ws[f'{XLSX_REQUEST_DATE}{row.row}'].value.strftime("%d.%m.%Y")) \
        if isinstance(ws[f'{XLSX_REQUEST_DATE}{row.row}'].value, datetime) \
           or isinstance(ws[f'{XLSX_REQUEST_DATE}{row.row}'].value, date) else ws[f'{XLSX_REQUEST_DATE}{row.row}'].value


def main_insert_and_sort_xlsx(data):
    filename = XLSX_FILE_STAVKA if data['postanovlenie'] == 'П' else XLSX_FILE_EDA
    document_name = f'../{filename}{XLSX_FILE_SUFFIX}'
    document = load_workbook(document_name)
    ws = document.active
    count_c = 1
    if "comment" not in data:
        comment = 'Соответствует'
    else:
        comment = ''
        for c in data.get('comment', '').split('\n'):
            comment += f'{count_c}. {c}\n'
            count_c += 1
    count = 0
    new_row = [data['name'], data['inn'], f'{data["postanovlenie"]}-{data["number"]}',
               data["request_date"], comment, data['summ'], data.get('yandex'), data.get('delivery'),
               data['ispolnitel'], data["request_date"]]
    rows = [[count, *new_row], ]
    begin_row = 3
    not_empty_row = 0
    for row in ws[XLSX_PP]:
        if row.row < 4:
            not_empty_row = row.row
            continue
        elif row.value:
            not_empty_row = row.row
            count += 1
            continue
        else:
            if row.row == not_empty_row + 1:
                begin_row = row.row
                count += 1
                rows = [[count, *new_row], ]
            else:
                break
    sorted_rows = sorted(rows, key=lambda r: r[3])
    for row in range(len(sorted_rows)):
        insert_values_in_row(ws, sorted_rows[row], begin_row + row)
    document.save(document_name)
